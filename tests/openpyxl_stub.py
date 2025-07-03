import sys
import types


def ensure_openpyxl_stub():
    """Install a lightweight openpyxl stub if the real package is missing."""
    try:
        import openpyxl  # noqa: F401
        return
    except Exception:
        pass

    openpyxl = types.ModuleType("openpyxl")
    openpyxl.load_workbook = lambda *a, **k: None

    styles_mod = types.ModuleType("openpyxl.styles")
    styles_mod.PatternFill = lambda **kw: None

    class Alignment:
        def __init__(self, **kwargs):
            pass

    styles_mod.Alignment = Alignment

    class Font:
        def __init__(self, **kwargs):
            pass

    styles_mod.Font = Font

    openpyxl.styles = styles_mod

    formatting_rule_mod = types.ModuleType("openpyxl.formatting.rule")

    class FormulaRule:
        def __init__(self, *a, **k):
            pass

    formatting_rule_mod.FormulaRule = FormulaRule
    sys.modules.setdefault("openpyxl.formatting.rule", formatting_rule_mod)

    utils_mod = types.ModuleType("openpyxl.utils")
    utils_mod.get_column_letter = lambda x: "A"
    ex_mod = types.ModuleType("openpyxl.utils.exceptions")
    ex_mod.InvalidFileException = Exception
    utils_mod.exceptions = ex_mod
    openpyxl.utils = utils_mod

    sys.modules.setdefault("openpyxl", openpyxl)
    sys.modules.setdefault("openpyxl.styles", styles_mod)
    sys.modules.setdefault("openpyxl.utils", utils_mod)
    sys.modules.setdefault("openpyxl.utils.exceptions", ex_mod)
