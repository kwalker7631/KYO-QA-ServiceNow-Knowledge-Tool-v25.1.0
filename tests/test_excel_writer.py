import sys
from pathlib import Path

sys.path.append(str(Path(__file__).resolve().parents[1]))

import types
import pytest
try:
    import openpyxl
except ModuleNotFoundError:
    openpyxl = types.ModuleType('openpyxl')
    class DummyWB:
        def __init__(self):
            self.active = types.SimpleNamespace()
        def save(self, *a, **k):
            pass
    openpyxl.Workbook = DummyWB
    def load_workbook(*a, **k):
        return DummyWB()
    openpyxl.load_workbook = load_workbook
    sys.modules['openpyxl'] = openpyxl

from excel_generator import ExcelWriter, NEEDS_REVIEW_FILL, FAILED_FILL
load_workbook = openpyxl.load_workbook


def test_conditional_row_fill(tmp_path):
    file_path = tmp_path / "out.xlsx"
    headers = ["processing_status", "value"]
    writer = ExcelWriter(str(file_path), headers)
    writer.add_row({"processing_status": "Needs Review", "value": "a"})
    writer.add_row({"processing_status": "Failed", "value": "b"})
    writer.save()

    wb = load_workbook(file_path)
    sheet = wb.active
    assert sheet["A2"].fill.start_color.rgb == NEEDS_REVIEW_FILL.start_color.rgb
    assert sheet["A3"].fill.start_color.rgb == FAILED_FILL.start_color.rgb



