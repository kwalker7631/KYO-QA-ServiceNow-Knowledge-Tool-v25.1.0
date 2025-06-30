import sys
from pathlib import Path
from types import SimpleNamespace
from queue import Queue

sys.path.append(str(Path(__file__).resolve().parents[1]))

import types
import data_harvesters

try:
    import openpyxl  # type: ignore
except ModuleNotFoundError:  # pragma: no cover - minimal stub for openpyxl
    openpyxl = types.ModuleType('openpyxl')
    _store = {}

    class _Cell:
        def __init__(self, value=None):
            self.value = value
            self.fill = None
            self.alignment = None

    class _Sheet:
        def __init__(self, rows=None):
            self.rows = [[_Cell(v) for v in row] for row in (rows or [])]
            self.column_dimensions = {}

        def __getitem__(self, idx):
            if isinstance(idx, str):
                col = ord(idx[0].upper()) - 64
                row = int(idx[1:])
                return self.cell(row, col)
            return self.rows[idx-1]

        def cell(self, row, column):
            while len(self.rows) < row:
                self.rows.append([_Cell() for _ in range(len(self.rows[0]))])
            while len(self.rows[row-1]) < column:
                self.rows[row-1].append(_Cell())
            return self.rows[row-1][column-1]

        def append(self, vals):
            self.rows.append([_Cell(v) for v in vals])

        def iter_rows(self, min_row=1):
            for r in self.rows[min_row-1:]:
                yield r

        @property
        def columns(self):
            for i in range(len(self.rows[0])):
                yield [row[i] if i < len(row) else _Cell() for row in self.rows]

        @property
        def max_row(self):
            return len(self.rows)

    class _Workbook:
        def __init__(self):
            self.active = _Sheet()

        def save(self, path):
            _store[str(path)] = self
            Path(path).touch()

    def Workbook():
        return _Workbook()

    def load_workbook(path):
        return _store.get(str(path), next(iter(_store.values())))

    class styles:
        class PatternFill:
            def __init__(self, *a, **k):
                pass
        class Alignment:
            def __init__(self, *a, **k):
                pass

    class utils:
        @staticmethod
        def get_column_letter(i):
            return chr(64 + i)

    openpyxl.Workbook = Workbook
    openpyxl.load_workbook = load_workbook
    openpyxl.styles = styles
    openpyxl.utils = utils
    sys.modules['openpyxl'] = openpyxl
    sys.modules['openpyxl.styles'] = styles
    sys.modules['openpyxl.utils'] = utils

import processing_engine


def test_run_processing_job_writes_meta(monkeypatch, tmp_path):
    # create basic Excel with required headers
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.append(["Short description", "Meta", "Author", "Processing Status"])
    sheet.append(["file.pdf", "", "", ""])
    input_excel = tmp_path / "input.xlsx"
    wb.save(input_excel)

    # dummy pdf path
    pdf_path = tmp_path / "file.pdf"
    pdf_path.write_text("dummy")

    dummy_result = {
        "filename": "file.pdf",
        "models": "ECOSYS M5526",
        "author": "",
        "status": "Pass",
        "ocr_used": False,
        "short_description": "file",
    }

    monkeypatch.setattr(processing_engine, "process_single_pdf", lambda p, q, ignore_cache=False: dummy_result)
    monkeypatch.setattr(processing_engine, "clear_review_folder", lambda: None)
    monkeypatch.setattr(processing_engine, "OUTPUT_DIR", tmp_path)
    monkeypatch.setattr(processing_engine, "PDF_TXT_DIR", tmp_path)
    monkeypatch.setattr(processing_engine, "is_file_locked", lambda p: False)
    monkeypatch.setattr(data_harvesters, "get_combined_patterns", lambda n, d: d)

    job_info = {"excel_path": str(input_excel), "input_path": [pdf_path]}
    progress_queue = Queue()
    cancel = SimpleNamespace(is_set=lambda: False)

    processing_engine.run_processing_job(job_info, progress_queue, cancel)

    output_files = list(tmp_path.glob('cloned_*.xlsx'))
    assert output_files
    out_wb = openpyxl.load_workbook(output_files[0])
    out_sheet = out_wb.active
    assert out_sheet["B2"].value == "ECOSYS M5526"
