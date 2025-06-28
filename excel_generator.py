# KYO QA ServiceNow Excel Generator - REFACTORED FOR SAFETY
from version import VERSION
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
import re
from logging_utils import setup_logger, log_info, log_error
from custom_exceptions import ExcelGenerationError

logger = setup_logger("excel_generator")

ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010\013\014\016-\037]')
NEEDS_REVIEW_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
OCR_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
FAILED_FILL = PatternFill(start_color="9C0006", end_color="9C0006", fill_type="solid")

DEFAULT_TEMPLATE_HEADERS = [
    'Active', 'Article type', 'Author', 'Category(category)', 'Configuration item',
    'Confidence', 'Description', 'Attachment link', 'Disable commenting',
    'Disable suggesting', 'Display attachments', 'Flagged', 'Governance',
    'Category(kb_category)', 'Knowledge base', 'Meta', 'Meta Description',
    'Ownership Group', 'Published', 'Scheduled publish date', 'Short description',
    'Article body', 'Topic', 'Problem Code', 'models', 'Ticket#', 'Valid to',
    'View as allowed', 'Wiki', 'Sys ID', 'file_name', 'Change Type', 'Revision'
]

def sanitize_for_excel(value):
    if isinstance(value, str): return ILLEGAL_CHARACTERS_RE.sub('', value)
    return value

def apply_styles_to_row(row, fill_color=None):
    for cell in row:
        cell.alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')
        if fill_color:
            cell.fill = fill_color

def finalize_styles(worksheet):
    header_font = Font(bold=True)
    for cell in worksheet[1]:
        cell.font = header_font
    for column_cells in worksheet.columns:
        try:
            max_length = max(len(str(cell.value)) for cell in column_cells if cell.value)
            worksheet.column_dimensions[column_cells[0].column_letter].width = min((max_length + 2), 70)
        except (ValueError, TypeError):
            continue
    log_info(logger, "Final worksheet styling complete.")

class ExcelWriter:
    def __init__(self, filepath, headers):
        self.filepath = filepath
        self.headers = [h.strip() for h in headers]  # Normalize whitespace
        self.header_map = {h.lower(): h for h in self.headers}  # Case-insensitive map
        self.workbook = openpyxl.Workbook()
        self.sheet = self.workbook.active
        self.sheet.title = "ServiceNow Import"
        self.sheet.append(self.headers)
        log_info(logger, f"ExcelWriter initialized for {filepath}")

    def add_row(self, data_dict):
        row_data = []
        for header in self.headers:
            value = ""
            for key in (header, header.lower(), header.strip(), header.replace(" ", "_")):
                if key in data_dict:
                    value = sanitize_for_excel(data_dict[key])
                    break
            row_data.append(value)

        self.sheet.append(row_data)

        status = data_dict.get('processing_status', 'Success')
        fill_color = None
        if status == 'Needs Review': fill_color = NEEDS_REVIEW_FILL
        elif status == 'OCR Required': fill_color = OCR_FILL
        elif status == 'Failed': fill_color = FAILED_FILL

        apply_styles_to_row(self.sheet[self.sheet.max_row], fill_color)

    def save(self):
        try:
            finalize_styles(self.sheet)
        except Exception as e:
            log_error(logger, f"Failed applying final Excel styles: {e}")

        try:
            self.workbook.save(self.filepath)
            log_info(logger, f"Successfully saved Excel file to {self.filepath}")
        except Exception as e:
            log_error(logger, f"Failed to save the final Excel file: {e}")
            raise ExcelGenerationError(f"Could not save Excel file: {e}")
